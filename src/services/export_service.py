"""
Export service for generating reports in different formats
"""
import json
import pandas as pd
from io import BytesIO
from datetime import datetime
from typing import List
from src.models.analysis import DomainResult

class ExportService:
    def __init__(self):
        pass
    
    def export_to_excel(self, results: List[DomainResult], task_name: str) -> bytes:
        """Export results to Excel format with formatting"""
        try:
            # Prepare data
            data = []
            for result in results:
                row = {
                    'Domain': result.domain_name,
                    'Has Snapshot': result.has_snapshot,
                    'Total Snapshots': result.total_snapshots,
                    'Timemap Count': result.timemap_count,
                    'First Snapshot': result.first_snapshot.strftime('%Y-%m-%d') if result.first_snapshot else '',
                    'Last Snapshot': result.last_snapshot.strftime('%Y-%m-%d') if result.last_snapshot else '',
                    'Avg Interval (days)': result.avg_interval_days,
                    'Max Gap (days)': result.max_gap_days,
                    'Years Covered': result.years_covered,
                    'Unique Versions': result.unique_versions,
                    'Is Good': result.is_good,
                    'Recommended': result.recommended,
                    'AI Category': result.ai_category,
                    'AI Confidence': result.ai_confidence,
                    'AI Description': result.ai_description,
                    'Assessment Score': result.assessment_score,
                    'Analysis Time (sec)': result.analysis_time_sec,
                    'Selected': result.is_selected
                }
                data.append(row)
            
            # Create DataFrame
            df = pd.DataFrame(data)
            
            # Create Excel file in memory
            output = BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='Domain Analysis', index=False)
                
                # Get workbook and worksheet for formatting
                workbook = writer.book
                worksheet = writer.sheets['Domain Analysis']
                
                # Apply formatting
                self._format_excel_worksheet(workbook, worksheet, df)
            
            output.seek(0)
            return output.getvalue()
            
        except Exception as e:
            raise Exception(f"Failed to export to Excel: {str(e)}")
    
    def export_to_csv(self, results: List[DomainResult]) -> bytes:
        """Export results to CSV format"""
        try:
            # Prepare data
            data = []
            for result in results:
                row = {
                    'domain': result.domain_name,
                    'has_snapshot': result.has_snapshot,
                    'total_snapshots': result.total_snapshots,
                    'timemap_count': result.timemap_count,
                    'first_snapshot': result.first_snapshot.isoformat() if result.first_snapshot else '',
                    'last_snapshot': result.last_snapshot.isoformat() if result.last_snapshot else '',
                    'avg_interval_days': result.avg_interval_days,
                    'max_gap_days': result.max_gap_days,
                    'years_covered': result.years_covered,
                    'unique_versions': result.unique_versions,
                    'is_good': result.is_good,
                    'recommended': result.recommended,
                    'ai_category': result.ai_category,
                    'ai_confidence': result.ai_confidence,
                    'ai_description': result.ai_description,
                    'assessment_score': result.assessment_score,
                    'analysis_time_sec': result.analysis_time_sec,
                    'is_selected': result.is_selected
                }
                data.append(row)
            
            # Create DataFrame and export to CSV
            df = pd.DataFrame(data)
            output = BytesIO()
            df.to_csv(output, index=False, encoding='utf-8')
            output.seek(0)
            return output.getvalue()
            
        except Exception as e:
            raise Exception(f"Failed to export to CSV: {str(e)}")
    
    def export_to_json(self, results: List[DomainResult]) -> bytes:
        """Export results to JSON format"""
        try:
            data = {
                'export_info': {
                    'timestamp': datetime.utcnow().isoformat(),
                    'total_domains': len(results),
                    'format': 'json'
                },
                'domains': [result.to_dict() for result in results]
            }
            
            json_str = json.dumps(data, indent=2, ensure_ascii=False)
            return json_str.encode('utf-8')
            
        except Exception as e:
            raise Exception(f"Failed to export to JSON: {str(e)}")
    
    def _format_excel_worksheet(self, workbook, worksheet, df):
        """Apply formatting to Excel worksheet"""
        try:
            from openpyxl.styles import PatternFill, Font, Alignment
            from openpyxl.formatting.rule import FormulaRule
            
            # Header formatting
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            
            for cell in worksheet[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            # Auto-adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
            
            # Conditional formatting for recommended domains
            green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            orange_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
            
            # Find column indices
            recommended_col = None
            last_snapshot_col = None
            
            for idx, col in enumerate(df.columns, 1):
                if col == 'Recommended':
                    recommended_col = idx
                elif col == 'Last Snapshot':
                    last_snapshot_col = idx
            
            # Apply conditional formatting
            if recommended_col:
                # Green for recommended domains
                worksheet.conditional_formatting.add(
                    f"A2:Z{len(df) + 1}",
                    FormulaRule(
                        formula=[f"${chr(64 + recommended_col)}2=TRUE"],
                        fill=green_fill
                    )
                )
            
            if last_snapshot_col:
                # Orange for recent snapshots (current year)
                current_year = datetime.now().year
                worksheet.conditional_formatting.add(
                    f"A2:Z{len(df) + 1}",
                    FormulaRule(
                        formula=[f"YEAR(${chr(64 + last_snapshot_col)}2)={current_year}"],
                        fill=orange_fill
                    )
                )
            
        except Exception as e:
            print(f"Warning: Could not apply Excel formatting: {str(e)}")
            # Continue without formatting if there's an error

